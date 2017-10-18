import os
import sqlite3 as sqlite
import openpyxl
import tkinter
from tkinter import ttk, filedialog, messagebox #separate imports needed due to tkinter idiosyncrasies

### local objects
from classes import stdevs, meanw, stdevw

### defines a new form used to select data the user wants to export to a list and then export it
class ExportForm:
    def __init__(self, parent, child, RDpath, connection):
        cframe = tkinter.Frame(child)
        cframe.grid()
        child.title("AIMRD Export")
        self.valueCat = []
        self.valueData = []
        self.valueScale = []
        self.connection = connection

        self.style = ttk.Style()
        self.style.configure("TButton", padding=6, relief="flat", background="#ccc", width=20)

        self.lblExport = ttk.Label(child, text="Export:")
        self.lblExport.grid(row=0, column = 0, columnspan =3, sticky = "W")

        self.lstCategory = tkinter.Listbox(child, selectmode = tkinter.EXTENDED, exportselection =0)
        self.lstCategory.grid(row=1, column = 0)

        self.lstDataType = tkinter.Listbox(child, selectmode = tkinter.EXTENDED, exportselection =0)
        self.lstDataType.grid(row=1, column = 1)

        self.lstScale = tkinter.Listbox(child, selectmode = tkinter.EXTENDED, exportselection =0)
        self.lstScale.grid(row=1, column = 2)

        self.btnSelectAll = ttk.Button(child, text='Select All', style="TButton", command=self.selectall)
        self.btnSelectAll.grid(row=2, column = 0)

        self.btnClear = ttk.Button(child, text='Clear Selection', style="TButton", command=self.clearall)
        self.btnClear.grid(row=2, column = 1)

        self.btnExport = ttk.Button(child, text='Export Selection', style="TButton", command=self.exportselection)
        self.btnExport.grid(row=2, column = 2)

        result = connection.execute("SELECT Category FROM Exports_All GROUP BY Category ORDER BY Category;")
        for row in result:
            # print(row)
            self.lstCategory.insert(tkinter.END, row['Category'])

        ### each of these 'onselect_' events defines what happens to the form when a specific widget is selected/changes selection.
        def onselect_Category(evt):
            self.lstDataType.delete(0, tkinter.END)
            self.lstScale.delete(0, tkinter.END)
            w = evt.widget
            c = w.curselection()
            value = []
            li = len(c)
            for i in range(0, li):
                value.append(w.get(c[i]))
            #print(value)
            self.valueCat = value

            ### this sql constuctor appears multiple times in this class. The {!s} and .format() construction allows for a
            ### variable number of ? to be inserted into the SQL for later parameter substitution. In this case it is
            ### constructing a "WHERE IN ('some', 'values', 'here')" clause with the values coming from the selected elements of the
            ### Category listbox.
            sql = "SELECT DataType FROM Exports_All WHERE Category IN ({!s}) GROUP BY DataType ORDER BY DataType;"
            sql = sql.format(','.join('?'*len(self.valueCat))) 
            #print(sql)
            result = connection.execute(sql, self.valueCat)
            for row in result:
                self.lstDataType.insert(tkinter.END, row['DataType'])

        def onselect_DataType(evt):
            self.lstScale.delete(0, tkinter.END)
            w = evt.widget
            c = w.curselection()
            value = []
            li = len(c)
            for i in range(0, li):
                value.append(w.get(c[i]))
            #print(value)
            self.valueData = value
            sql = "SELECT Scale FROM Exports_All WHERE DataType IN ({!s}) AND Category IN ({!s}) GROUP BY Scale ORDER BY Scale;"
            sql = sql.format(','.join('?'*len(self.valueData)), ",".join('?'*len(self.valueCat)))
            #print(sql)
            result = connection.execute(sql, self.valueData + self.valueCat)
            for row in result:
                #print(row)
                self.lstScale.insert(tkinter.END, row['Scale'])

        def onselect_Scale(evt):
            w = evt.widget
            c = w.curselection()
            value = []
            li = len(c)
            for i in range(0, li):
                value.append(w.get(c[i]))
            #print(value)
            self.valueScale = value

        ### this statements bind the 'onselect_' functions with the actual action of selecting an item from a listbox.
        self.lstCategory.bind('<<ListboxSelect>>', onselect_Category)
        self.lstDataType.bind('<<ListboxSelect>>', onselect_DataType)
        self.lstScale.bind('<<ListboxSelect>>', onselect_Scale)

    def selectall(self):
        self.lstCategory.select_set(0, tkinter.END)
        c = self.lstCategory.curselection()
        value = []
        li = len(c)
        for i in range(0, li):
            value.append(self.lstCategory.get(c[i]))
        self.valueCat = value

        result = self.connection.execute("SELECT DataType FROM Exports_All GROUP BY DataType ORDER BY DataType;")
        for row in result:
            self.lstDataType.insert(tkinter.END, row['DataType'])
        self.lstDataType.select_set(0, tkinter.END)
        c = self.lstDataType.curselection()
        value = []
        li = len(c)
        for i in range(0, li):
            value.append(self.lstDataType.get(c[i]))
        self.valueData = value

        result = self.connection.execute("SELECT Scale FROM Exports_All GROUP BY Scale ORDER BY Scale;")
        for row in result:
            self.lstScale.insert(tkinter.END, row['Scale'])
        self.lstScale.select_set(0, tkinter.END)
        c = self.lstScale.curselection()
        value = []
        li = len(c)
        for i in range(0, li):
            value.append(self.lstScale.get(c[i]))
        self.valueScale = value

    def clearall(self):
        self.lstDataType.delete(0, tkinter.END)
        self.lstScale.delete(0, tkinter.END)
        self.lstCategory.selection_clear(0, tkinter.END)

    def exportselection(self):        
        self.connection.enable_load_extension(True)
        self.connection.execute("SELECT load_extension('mod_spatialite')")
        sql = """SELECT ObjectName, ExportName
                   FROM Exports_All
                  WHERE Category IN ({!s}) AND DataType IN ({!s}) AND Scale IN ({!s}) 
                  ORDER BY Category, Scale, DataType, ExportName;"""
        sql = sql.format(','.join('?'*len(self.valueCat)), ','.join('?'*len(self.valueData)), ','.join('?'*len(self.valueScale)))
        result = self.connection.execute(sql, self.valueCat + self.valueData + self.valueScale)
        path = tkinter.filedialog.asksaveasfilename(defaultextension = ".xlsx", title="Choose filename for export:", filetypes=(("Excel files", "*.xlsx"),("All files", "*.*")))
        if os.path.isfile(path):
            os.remove(path)
        if path:
            try:
                workbook = openpyxl.Workbook()
                exportEmpty = True # this checks to see if the user wants to export empty views/tables
                asked = False # this keeps track of whether the user was asked to export blanks or not
                for row in result:
                    if os.path.isfile(path):    
                        workbook = openpyxl.load_workbook(filename = path)
                    else:
                        workbook = openpyxl.Workbook()
                    self.lblExport['text'] = "                                                                                          "
                    self.lblExport.update_idletasks()
                    self.lblExport['text'] = "Exporting " + row["ExportName"] + "..."
                    self.lblExport.update_idletasks()

                    ### this section was an attmept to use temporary tables in order to try and deal with MemoryError exceptions.
                    ### unfortunately the loop was attempting to drop temp tables while they were still in use by other parts of the loop.
                    ### may still work with extra tweaking.
                    #self.connection.execute("DROP TABLE IF EXISTS OutTable;")
                    #self.connection.execute("CREATE TEMPORARY TABLE OutTable AS SELECT * FROM {!s};".format(row["ObjectName"]))
                    #obj = 'OutTable'
                    
                    ### figures out how many rows are in a query to be exported and then creates the relevant recordset
                    ### unfortunately due to the size and complexity of some views this has produced MemoryError exceptions in the past.
                    ### still needs to be fixed if possible.
                    obj = row["ObjectName"]
                    rowcount = self.connection.execute("SELECT Count(*) FROM {!s};".format(obj)).fetchone()[0]
                    result2 = self.connection.execute("SELECT * FROM {!s};".format(obj))

                    if rowcount == 0:
                        if not asked:
                            exportEmpty = tkinter.messagebox.askyesno("Export Blanks", "Export blank results?")
                            asked = True
                    if rowcount > 0 or (rowcount == 0 and exportEmpty):          
                        print("Exporting",row["ExportName"],"...")
                        if os.path.isfile(path):
                            worksheet = workbook.create_sheet(title=row["ExportName"])
                        else:
                            worksheet = workbook.active
                            worksheet.title = row["ExportName"]
                        colnames = [desc[0] for desc in result2.description]
                        worksheet.append(colnames)
                        for row2 in result2:
                            d = dict(row2)
                            for key, value in d.items(): #necessary in order to tell excel that strings starting with '=' are not formulas. Not an ideal result, could use tweaking.
                                if isinstance(value, str):
                                    if len(value) > 0:
                                        if(value[0]=='='):
                                            d[key]= ''.join(("'",value))
                            worksheet.append(list(d.values()))
                        workbook.save(path)
                    else:
                        print("Skipping",row["ExportName"],"...")
                print("Finished Exporting.")
                tkinter.messagebox.showinfo("Success", "Export complete.")
                self.lblExport['text'] = "Export:"
                self.lblExport.update_idletasks()
                self.clearall()

            except PermissionError:
                tkinter.messagebox.showinfo("Error", "Could not access" + os.linesep + path + os.linesep +"File may be in use.")

def Export(RDpath, form = None):
    dirpath = os.path.dirname(RDpath)
    dbname = os.path.basename(RDpath)
    conn = sqlite.connect(RDpath)
    conn.row_factory = sqlite.Row
    conn.create_aggregate("stdev", 1, stdevs)
    conn.create_aggregate("meanw", 2, meanw)
    conn.create_aggregate("stdevw", 2, stdevw)
    child = tkinter.Tk()
    cf = ExportForm(form, child, RDpath, conn)
    child.mainloop()
    conn.close()
    return;



